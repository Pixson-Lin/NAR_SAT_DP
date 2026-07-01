"""GNSS 解析結果輸出為 CSV 與 Excel。"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .gnss_parser import MAX_SIGNALS, GnssRow

_FIXED_HEADERS = [
    "hostname",
    "Control",
    "Elev. Mask Angle",
    "Used Satellite(Control)",
    "Constellation",
    "Used Satellite(Constellation)",
]


def build_headers(include_source: bool = False) -> list[str]:
    headers = list(_FIXED_HEADERS)
    for i in range(1, MAX_SIGNALS + 1):
        headers.append(f"A signal {i}")
    for i in range(1, MAX_SIGNALS + 1):
        headers.append(f"B signal {i}")
    headers.append("script_begin_time")
    if include_source:
        headers.extend(["source_archive", "source_txt_path"])
    return headers


def write_gnss_csv(
    path: Path,
    rows: list[GnssRow],
    encoding: str = "utf-8-sig",
    include_source: bool = False,
) -> None:
    headers = build_headers(include_source=include_source)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding=encoding, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.to_flat_dict().get(h, "") for h in headers})


def write_gnss_xlsx(
    path: Path,
    rows: list[GnssRow],
    include_source: bool = False,
) -> None:
    headers = build_headers(include_source=include_source)
    wb = Workbook()
    ws = wb.active
    ws.title = "GNSS"

    fixed_count = len(_FIXED_HEADERS)
    hostname_col = 1
    a_start = fixed_count + 1
    a_end = a_start + MAX_SIGNALS - 1
    b_start = a_end + 1
    b_end = b_start + MAX_SIGNALS - 1
    time_col = b_end + 1
    source_archive_col = time_col + 1 if include_source else None
    source_txt_col = time_col + 2 if include_source else None
    last_col = source_txt_col if include_source else time_col

    col = 1
    for title in _FIXED_HEADERS:
        ws.cell(1, col, title)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

    ws.cell(1, a_start, f"A signal 1...{MAX_SIGNALS}")
    ws.merge_cells(start_row=1, start_column=a_start, end_row=1, end_column=a_end)
    for i in range(MAX_SIGNALS):
        ws.cell(2, a_start + i, str(i + 1))

    ws.cell(1, b_start, f"B signal 1...{MAX_SIGNALS}")
    ws.merge_cells(start_row=1, start_column=b_start, end_row=1, end_column=b_end)
    for i in range(MAX_SIGNALS):
        ws.cell(2, b_start + i, str(i + 1))

    ws.cell(1, time_col, "script_begin_time")
    ws.merge_cells(start_row=1, start_column=time_col, end_row=2, end_column=time_col)

    if include_source:
        ws.cell(1, source_archive_col, "source_archive")
        ws.merge_cells(
            start_row=1,
            start_column=source_archive_col,
            end_row=2,
            end_column=source_archive_col,
        )
        ws.cell(1, source_txt_col, "source_txt_path")
        ws.merge_cells(
            start_row=1,
            start_column=source_txt_col,
            end_row=2,
            end_column=source_txt_col,
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[2]:
        cell.alignment = Alignment(horizontal="center")

    data_start = 3
    for idx, row in enumerate(rows):
        r = data_start + idx
        flat = row.to_flat_dict()
        for c, header in enumerate(headers, start=1):
            ws.cell(r, c, flat.get(header, ""))

    merge_cols = [hostname_col, time_col]
    if include_source:
        merge_cols.extend([source_archive_col, source_txt_col])

    ne_start = data_start
    while ne_start < data_start + len(rows):
        ne_end = min(ne_start + 1, data_start + len(rows) - 1)
        if ne_end > ne_start:
            for merge_col in merge_cols:
                ws.merge_cells(
                    start_row=ne_start,
                    start_column=merge_col,
                    end_row=ne_end,
                    end_column=merge_col,
                )
        ne_start += 2

    for c in range(1, last_col + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = 14 if c >= a_start else 18

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
